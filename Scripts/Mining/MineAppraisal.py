import pandas as pd
import openpyxl, re, logging
from openpyxl import load_workbook
from datetime import datetime

# Configure logging
logging.basicConfig(filename='../../files/Logs/MineAppraisal.log', format='%(message)s', level=logging.INFO)
logging.info(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

def find_gdv(file_path):
    # Load the Excel file
    wb = load_workbook(file_path, data_only=True)

    # Set up 3 cell highlight colours
    colour_1 = openpyxl.styles.colors.Color(rgb='00CCFFFF')
    colour_2 = openpyxl.styles.colors.Color(rgb='0033CCCC')
    colour_3 = openpyxl.styles.colors.Color(rgb='0099CC00')
    fill_1 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colour_1)
    fill_2 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colour_2)
    fill_3 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colour_3)


    # Search for GDV column names or acronyms
    gdv_columns = ['Purchase Price', 'Total Build Loan', 'S106', 'Bank Legal Fees', 'Address', 'Commercial',
                   'demoli', 'stamp', 'legal', 'agent', 'residential', 'common', 'contingency',
                   ]

    # Iterate through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        first_postcode = False
        # Iterate through each cell in the sheet
        for row in sheet.iter_rows():
            for cell in row:

                # Find the first postcode in the file (assume this is postcode of the project)
                if not first_postcode:
                    postcode = re.findall(r'[A-Z]{1,2}[0-9R][0-9A-Z]? [0-9][A-Z]{2}', str(cell.value))
                    if postcode:
                        print('First Postcode: ', postcode[0], '\n')
                        logging.info('First Postcode: '+ postcode[0] + '\n')
                        first_postcode = True
                        cell.fill = fill_2

                # Check if the cell value contains a keyword
                for keyword in gdv_columns:
                    if keyword.lower() in str(cell.value).lower():

                        # Highlight cells that contain keywords
                        cell.fill = fill_1

                        # Find values in keyword's right hand neighbours
                        for i in range(1,4):
                            if cell.offset(column=i).value is not None:
                                print(cell.value, ': ', cell.offset(column=i).value)
                                print('    ',cell.offset(column=i))
                                cell.offset(column=i).fill = fill_2
                                if type(cell.offset(column=i).value) != str:

                                    # Flag value as percentage if 0<x<1
                                    if 0 < cell.offset(column=i).value < 1  :
                                        print("     % Percentage (Assumed)", '\n')
                                        cell.offset(column=i).fill = fill_3
                                        # break
                                    if cell.offset(column=i + 1).value == 0: print()

                                    # Flag value as not total
                                    if cell.offset(column=i + 1).value is not None and cell.offset(column=i).value>1 :
                                        print(     "1/x Unit Cost (Assumed)", '\n')

                                    #Flag value as total
                                    if cell.offset(column=i + 1).value is None and cell.offset(column=i + 2).value is None:
                                        print("     * Total (Assumed)", '\n')
                                        cell.offset(column=i).fill = fill_2
                                        # break
                                else: print()

    # Save highlighted file
    new_file = file_path.replace('.xlsx', '') + ' Highlighted.xlsx'
    wb.save(new_file)

# Provide the file path of your Excel file
full_file_path = "/Users/ruben/PycharmProjects/PDFExtract/files/Development Appraisal.xlsx"

# Call the function to find and print the found values
print()
find_gdv(full_file_path)
#
#
# v1
# import pandas as pd
#
# def find_gdv(file_path):
#     # Read the Excel file
#     df = pd.read_excel(file_path)
#
#     # Search for GDV column names or acronyms
#     gdv_columns = ['GDV', 'Gross Development Value']
#
#     # Iterate through each column and find GDV
#     for column in df.columns:
#         # Check if the column name contains GDV or its acronym
#         if any(keyword.lower() in column.lower() for keyword in gdv_columns):
#             # Iterate through each row in the column
#             for value in df[column]:
#                 # Check if the value is not empty
#                 if pd.notnull(value):
#                     # Print the GDV value
#                     print("GDV:", value)
#                     return
#
#     # If GDV is not found
#     print("GDV not found in the Excel file.")
#
# # Provide the file path of your Excel file
# file_path = "path/to/your/file.xlsx"
#
# # Call the function to find and print the GDV
# find_gdv(file_path)
#
#
# v2
# import pandas as pd
# from openpyxl import load_workbook
#
# def find_gdv(file_path):
#     # Load the Excel file
#     wb = load_workbook(file_path, read_only=True, data_only=True)
#
#     # Search for GDV column names or acronyms
#     gdv_columns = ['GDV', 'Gross Development Value']
#
#     # Iterate through each sheet in the workbook
#     for sheet_name in wb.sheetnames:
#         sheet = wb[sheet_name]
#
#         # Iterate through each cell in the sheet
#         for row in sheet.iter_rows():
#             for cell in row:
#                 # Check if the cell value contains GDV or its acronym
#                 if any(keyword.lower() in str(cell.value).lower() for keyword in gdv_columns):
#                     # Print the GDV value
#                     print("GDV:", cell.offset(column=1).value)
#                     return
#
#     # If GDV is not found
#     print("GDV not found in the Excel file.")
#
# # Provide the file path of your Excel file
# file_path = "path/to/your/file.xlsx"
#
# # Call the function to find and print the GDV
# find_gdv(file_path)
#
# v3
# import pandas as pd
#
# def find_gdv(file_path):
#     # Read the Excel file
#     df = pd.read_excel(file_path)
#
#     # Search for GDV column names or acronyms
#     gdv_columns = ['GDV', 'Gross Development Value']
#
#     # Iterate through each column and find GDV
#     for column in df.columns:
#         # Check if the column name contains GDV or its acronym
#         if any(keyword.lower() in column.lower() for keyword in gdv_columns):
#             # Iterate through each cell in the column
#             for index, value in df[column].items():
#                 # Check if the value is not empty
#                 if pd.notnull(value):
#                     # Check if the value is numeric
#                     if pd.to_numeric(value, errors='coerce') == value:
#                         # Print the GDV value
#                         print("GDV:", value)
#                         return
#                     else:
#                         # Check the next cells in the row for numeric values
#                         for next_value in df.iloc[index, :].values[index+1:]:
#                             if pd.notnull(next_value) and pd.to_numeric(next_value, errors='coerce') == next_value:
#                                 # Print the GDV value found in the next cell
#                                 print("GDV:", next_value)
#                                 return
#
#     # If GDV is not found
#     print("GDV not found in the Excel file.")
#
# # Provide the file path of your Excel file
# file_path = "path/to/your/file.xlsx"
#
# # Call the function to find and print the GDV
# find_gdv(file_path)
#
# v4
# import pandas as pd
#
# def find_gdv(file_path):
#     # Read the Excel file
#     df = pd.read_excel(file_path)
#
#     # Search for GDV column names or acronyms
#     gdv_columns = ['Purchase Price', 'Total Purchase Costs']
#
#     # Iterate through each column and find GDV
#     for column in df.columns:
#         # Check if the column name contains GDV or its acronym
#         if any(keyword.lower() in column.lower() for keyword in gdv_columns):
#             # Get the column index
#             column_index = df.columns.get_loc(column)
#             print('keyword found')
#
#             # Iterate through each row in the column
#             for index, value in enumerate(df.iloc[:, column_index]):
#                 # Check if the value is not empty
#                 if pd.notnull(value):
#                     # Check if the next row in the same column contains a numeric value
#                     next_row_value = df.iloc[index + 1, column_index]
#                     if pd.notnull(next_row_value) and isinstance(next_row_value, (int, float)):
#                         # Print the GDV value
#                         print("GDV:", next_row_value)
#                         return
#
#     # If GDV is not found
#     print("GDV not found in the Excel file.")
#
# # Provide the file path of your Excel file
# file_path = "/Users/ruben/PycharmProjects/PDFExtract/files/Development Appraisal.xlsx"
#
# # Call the function to find and print the GDV
# find_gdv(file_path)

# v5
# import pandas as pd
#
# def find_gdv(file_path):
#     # Read the Excel file
#     xls = pd.ExcelFile(file_path)
#
#     # Iterate through each sheet in the Excel file
#     for sheet_name in xls.sheet_names:
#         # Read the sheet into a DataFrame
#         df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
#
#         # Iterate through each cell in the DataFrame
#         for row in df.values:
#             for cell in row:
#                 # Check if the cell contains GDV or its acronym
#                 if isinstance(cell, str) and any(keyword.lower() in cell.lower() for keyword in ['GDV', 'Gross Development Value']):
#                     # Print the GDV value
#                     print("GDV:", cell)
#                     return
#
#     # If GDV is not found
#     print("GDV not found in the Excel file.")
#
# # Provide the file path of your Excel file
# file_path = "/Users/ruben/PycharmProjects/PDFExtract/files/Development Appraisal.xlsx"
#
# # Call the function to find and print the GDV
# find_gdv(file_path)

# v6
# import pandas as pd
#
# def find_values(file_path, keywords):
#     # Read the Excel file
#     df = pd.read_excel(file_path)
#
#     # Iterate through each keyword
#     for keyword in keywords:
#         # Search for columns containing the keyword
#         matching_columns = [column for column in df.columns if keyword.lower() in column.lower()]
#
#         # Iterate through each matching column
#         for column in matching_columns:
#             # Iterate through each row in the column
#             for value in df[column]:
#                 # Check if the value is not empty
#                 if pd.notnull(value):
#                     # Print the keyword and corresponding value
#                     print(keyword + ":", value)
#
# # Provide the file path of your Excel file
# file_path = "/Users/ruben/PycharmProjects/PDFExtract/files/Development Appraisal.xlsx"
#
# # Provide the keywords to search for
# keywords = ['GDV', 'Gross Development Value', 'Purchase Price']
#
# # Call the function to find and print the values
# find_values(file_path, keywords)