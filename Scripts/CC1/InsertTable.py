from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import re, string, docx

def column_to_number(column_letters):
    column_number = 0
    for letter in column_letters:
        if letter.upper() in string.ascii_uppercase:
            column_number = column_number * 26 + (ord(letter.upper()) - ord('A')) + 1
    return column_number

def get_table_dimensions(ref_str):
    # Split table ref into first and last number, first and last letter(s)
    start, end = ref_str.split(':')
    # half1 = split1[0]
    # half2 = split1[1]
    start_row = re.search('[0-9]+', start).group()
    end_row = re.search('[0-9]+', end).group()
    end_col = re.search('[A-Z]+', end).group()
    start_col = re.search('[A-Z]+', start).group()

    # print(n1, n2, l1, l2)

    # Get dimensions of table from start/end indices
    height = int(end_row) - int(start_row) + 1
    width = column_to_number(end_col) - column_to_number(start_col) + 1

    return width, height

def get_table_data(ref_str, cell_range, name, context):
    # Get dimensions of table
    width, height = get_table_dimensions(ref_str)
    df = pd.DataFrame(index=range(height), columns=range(width))
    # Get values of table as df
    for row in range(height):
        for col in range(width):
            df.iloc[row, col] = cell_range[row][col].value or ''

    df.columns = df.iloc[0]
    df = df[1:]
    split_dict = df.to_dict('split')
    data_dict = split_dict['data']

    for i in range(len(data_dict)):
        data_dict[i] = {'cols' : data_dict[i]}
    # print(data_dict)
    columns = split_dict['columns']
    # context = print_to_table(data_dict, columns, name, context)
    context[f'{name}_labels'] = df.columns.tolist()
    context[f'{name}_contents'] = data_dict

    return context

def insert_dataframe(df, template, num):
    my_context = {'table': df}
    template.render(my_context)
    new_file = "../TestTableInsert.docx"
    # print("Saving File")
    template.save(new_file)

    doc = docx.Document(new_file)

    # add a table to the end and create a reference variable
    t = doc.add_table(df.shape[0], df.shape[1])

    # add the rest of the data frame
    for i in range(df.shape[0]):
        for j in range(df.shape[-1]):
            t.cell(i, j).text = str(df.values[i, j])

    # save the doc
    doc.save(new_file)


def main(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Coversheet"]
    template = DocxTemplate("/Users/ruben/PycharmProjects/Shojin/files/InsertTableFiles/Sources/dynamic_table_tpl3.docx")
    index = 1
    context = {}
    for table in ws.tables.values():
        name = table.name
        ref = table.ref
        cell_range = ws[ref]
        context = get_table_data(ref, cell_range, name, context)
        index += 1
        # insert_dataframe(df, template, i)
    template.render(context)
    template.save(f"/Users/ruben/PycharmProjects/Shojin/files/InsertTableFiles/Generated/dynamic_table_out_whole.docx")



if __name__ == '__main__':
    file_path = "/Users/ruben/PycharmProjects/Shojin/files/InsertTableFiles/TestTables.xlsx"
    file_path = "/Users/ruben/PycharmProjects/Shojin/files/InsertTableFiles/TestTables.xlsx"
    main(file_path)


#    Dynamic Tables: https://medium.com/@lukas.forst/supercharge-your-word-reports-using-python-docx-template-4e9ebfc66b9e