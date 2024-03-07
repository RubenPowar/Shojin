import openpyxl, os
from openpyxl import load_workbook



def find_attributes():
#Timeline
    construction_months = 12
    refi_period = 6



#Sales
    gdv= 6733950
    sales_agents_fees = 101009
    sales_legal = 20202

#Ungeared
    #Acquisition
    purchase_price = 1500000
    stamp_duty = 75000
    acquisition_legal =11250
    acquisition_agents_fee = 0
    acquisition_valuation = 0
    dd_fees = 0
    acquisition_agent_fees = 0
    #Development
    residential_build_cost  =2493200
    commercial_space = 0
    common_areas = 160875
    contingency = 265
    professional_fees = 0
    demolition = 0
    monitoring = 0
    #Other
    s106 = 190262
    party_wall = 0
    pre_planning_costs = 0
    cil = 0
    marketing = 0

#Funding (has default values)
    #Net Facility
    acquisition_loan = 576250
    development_loan = 3390153
    s106_loan = 0
    marketing_loan = 0
    #Rolled Up Interest
    arrangement_fee = 2
    monitoring_fee = 15000
    committment_fee = 0
    interest_loan = 355000
    margin_base = 6
    base = 5.25
    #Serviced Interest
    serviced_margin_base = 0
    base_sonia = 0
    #Serviced fees
    serviced_dd_fee = 0
    serviced_monitoring = 0
    serviced_commitment_fee = 0
    exit_fee = 1

    write_timeline(construction_months, refi_period)
    write_sales(gdv, sales_agents_fees, sales_legal)
    write_acquisition(purchase_price, stamp_duty, acquisition_legal, acquisition_agents_fee,
                    acquisition_valuation, dd_fees, acquisition_agent_fees)
    write_development(residential_build_cost, commercial_space, common_areas, contingency,
                    professional_fees, demolition,monitoring)
    write_other(s106, party_wall, pre_planning_costs, cil, marketing)
    write_net_facility(acquisition_loan, development_loan, s106_loan, marketing_loan)
    write_rolled_up(arrangement_fee, monitoring_fee, committment_fee, interest_loan, margin_base, base)
    write_serviced_interest(serviced_margin_base, base_sonia)
    write_serviced_fees(serviced_dd_fee, serviced_monitoring, serviced_commitment_fee, exit_fee)




def write_timeline(cm, rp):
    ws['I1'] = cm
    ws['I2'] = rp

def write_sales(gdv, af, leg):
    ws['I4'] = gdv
    ws['I5'] = af
    ws['I6'] = leg

def write_acquisition(pp, sd, leg, af, val, dd, afs):
    ws['I8']  = pp
    ws['I9']  = sd
    ws['I10'] = leg
    ws['I11'] = af
    ws['I12'] = val
    ws['I13'] = dd
    ws['I14'] = afs

def write_development(rbc, cs, ca, con, pro, dem, mon):
    ws['I16'] = rbc
    ws['I17'] = cs
    ws['I18'] = ca
    ws['I19'] = con
    ws['I20'] = pro
    ws['I21'] = dem
    ws['I22'] = mon

def write_other(s1, pw, ppc, cil, mar):
    ws['I24'] = s1
    ws['I25'] = pw
    ws['I26'] = ppc
    ws['I27'] = cil
    ws['I29'] = mar

def write_net_facility(acq, dev, s1, mar):
    ws['I30'] = acq
    ws['I31'] = dev
    ws['I32'] = s1
    ws['I33'] = mar

def write_rolled_up(arr, mon, com, inl, mb, bas):
    ws['I35'] = arr
    ws['I36'] = mon
    ws['I37'] = com
    ws['I38'] = inl
    ws['I39'] = mb
    ws['I40'] = bas

def write_serviced_interest(mb, bas):
    ws['I42'] = mb
    ws['I43'] = bas

def write_serviced_fees(dd, mon, com, exf):
    ws['I45'] = dd
    ws['I46'] = mon
    ws['I47'] = com
    ws['I48'] = exf

def author(file):
    return file.properties.creator


for file in os.listdir('/Users/ruben/PycharmProjects/PDFExtract/venv/Files'):
    filename = os.fsdecode(file)
    print("file")

filename= 'Files/Appraisal.xlsx'

wb = load_workbook(filename=filename, data_only=true) #data_only=true ensures that formulae are converted to values maybe pandas does this by default
print('Author of \'' + filename + '\' is ' + author(wb))
ws = wb.worksheets[0]
find_attributes()

wb.save(file)